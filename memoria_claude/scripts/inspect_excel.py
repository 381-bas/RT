# -*- coding: utf-8 -*-
"""
inspect_excel.py -- inspeccion rapida y reusable de cualquier .xlsx/.xlsm, en vez de
escribir un script desechable cada vez (paso muy seguido en varias sesiones: inspect_
ligero.py, inspect_ligero2.py, debug_plantilla.py, dump_full_plantilla.py... todos
reinventaban lo mismo: leer headers, contar filas, filtrar por cliente).

Es SOLO LECTURA (openpyxl, read_only=True salvo que se pida --tablas) -- nunca escribe
nada, nunca arriesga el archivo. Ver memoria_claude/01_kernel.json.

Ejemplos:
    # Listar todas las hojas con sus dimensiones
    python inspect_excel.py archivo.xlsm

    # Ver headers + primeras filas de una hoja
    python inspect_excel.py archivo.xlsm --hoja OSA

    # Ver mas filas de muestra
    python inspect_excel.py archivo.xlsm --hoja MINUTA --filas 15

    # Contar valores de una columna (ej. cuantas filas por CLIENTE)
    python inspect_excel.py archivo.xlsm --hoja REPORTES --contar CLIENTE

    # Filtrar filas por un valor exacto de columna (ej. ver solo MORETTA en OSA)
    python inspect_excel.py archivo.xlsm --hoja OSA --filtro CLIENTE=MORETTA --filas 20

    # Listar las Tablas de Excel (ListObjects) de una hoja -- requiere read_only=False,
    # openpyxl no expone ListObjects en modo solo-lectura.
    python inspect_excel.py archivo.xlsm --hoja MINUTA --tablas
"""
import argparse
import sys
from collections import Counter

import openpyxl


def listar_hojas(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    print(f"Archivo: {path}")
    print(f"{'Hoja':<28} {'Filas':>8} {'Columnas':>10}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"{name:<28} {ws.max_row:>8} {ws.max_column:>10}")
    wb.close()


def leer_headers(ws):
    fila1 = next(ws.iter_rows(min_row=1, max_row=1))
    return [c.value for c in fila1]


def mostrar_hoja(path, hoja, n_filas):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    if hoja not in wb.sheetnames:
        print(f"ERROR: la hoja '{hoja}' no existe. Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[hoja]
    headers = leer_headers(ws)
    print(f"Hoja: {hoja} | filas: {ws.max_row} | columnas: {ws.max_column}")
    print(f"Headers: {headers}")
    print()
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=1 + n_filas)):
        print([c.value for c in row])
    wb.close()


def contar_columna(path, hoja, columna):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    ws = wb[hoja]
    headers = leer_headers(ws)
    if columna not in headers:
        print(f"ERROR: columna '{columna}' no existe en '{hoja}'. Columnas: {headers}")
        sys.exit(1)
    idx = headers.index(columna)
    contador = Counter()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        contador[row[idx].value] += 1
    print(f"Conteo de '{columna}' en '{hoja}' ({len(contador)} valores distintos):")
    for valor, n in contador.most_common(30):
        print(f"  {valor!r}: {n}")
    wb.close()


def filtrar_filas(path, hoja, filtro, n_filas):
    columna, _, valor = filtro.partition("=")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    ws = wb[hoja]
    headers = leer_headers(ws)
    if columna not in headers:
        print(f"ERROR: columna '{columna}' no existe en '{hoja}'. Columnas: {headers}")
        sys.exit(1)
    idx = headers.index(columna)
    print(f"Headers: {headers}")
    n_mostradas = 0
    n_total = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if str(row[idx].value) == valor:
            n_total += 1
            if n_mostradas < n_filas:
                print([c.value for c in row])
                n_mostradas += 1
    print(f"\n{n_total} fila(s) con {columna}={valor!r} (mostradas: {min(n_mostradas, n_filas)})")
    wb.close()


def listar_tablas(path, hoja):
    # ListObjects (Tablas de Excel) NO se exponen en modo read_only -- unica excepcion
    # a "siempre leer read_only=True" de este script.
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True, keep_vba=True)
    ws = wb[hoja]
    tablas = list(ws.tables.keys()) if hasattr(ws, "tables") else []
    print(f"Tablas (ListObjects) en '{hoja}': {tablas}")
    wb.close()


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("archivo", help="Ruta al .xlsx/.xlsm a inspeccionar")
    ap.add_argument("--hoja", help="Nombre de la hoja a inspeccionar (sin esto, lista todas)")
    ap.add_argument("--filas", type=int, default=5, help="Cantidad de filas de muestra (default 5)")
    ap.add_argument("--contar", metavar="COLUMNA", help="Cuenta valores distintos de esta columna")
    ap.add_argument("--filtro", metavar="COLUMNA=VALOR", help="Muestra solo filas que matcheen (comparacion exacta como texto)")
    ap.add_argument("--tablas", action="store_true", help="Lista las Tablas de Excel (ListObjects) de --hoja")
    args = ap.parse_args()

    if not args.hoja:
        listar_hojas(args.archivo)
        return

    if args.tablas:
        listar_tablas(args.archivo, args.hoja)
        return

    if args.contar:
        contar_columna(args.archivo, args.hoja, args.contar)
        return

    if args.filtro:
        filtrar_filas(args.archivo, args.hoja, args.filtro, args.filas)
        return

    mostrar_hoja(args.archivo, args.hoja, args.filas)


if __name__ == "__main__":
    main()
