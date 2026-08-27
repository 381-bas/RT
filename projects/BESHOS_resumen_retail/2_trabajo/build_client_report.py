# -*- coding: utf-8 -*-
"""
Genera un archivo NUEVO y autocontenido (no una copia/splice del archivo base)
a partir de los datos ya calculados en analisis_beshos_retail.json.

Por que un archivo nuevo y no insertar hojas en ANALISIS_BESHOS.xlsx:
ese libro es enorme (varias hojas de 50-80MB de XML) y la unica forma segura
de insertarle hojas es cirugia de texto sobre el XML -- funciono, pero cada
insercion nueva es una superficie de riesgo distinta (namespaces, posiciones,
relaciones) y Excel puede volver a marcarlo para reparar. Un archivo propio,
chico, armado 100% con openpyxl, no tiene ese riesgo: es OOXML valido por
construccion. Este es el archivo pensado para mostrarle al cliente.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.label import DataLabelList, DataLabel
from openpyxl.utils import get_column_letter

DATA_PATH = r"C:\Users\basti\Desktop\RT\CLIENTE_GG\analisis_beshos_retail.json"
OUT_PATH = r"C:\Users\basti\Desktop\RT\CLIENTE_GG\BESHOS - Resumen Retail (Cliente).xlsx"

data = json.load(open(DATA_PATH, encoding="utf-8"))
R = data["retailers"]
RETAILERS = ["WALMART", "JUMBO", "SANTA ISABEL", "SMU", "TOTTUS"]
LABELS = {"WALMART": "Walmart", "JUMBO": "Jumbo", "SANTA ISABEL": "Sta. Isabel",
          "SMU": "SMU", "TOTTUS": "Tottus"}
BASE_WEEK = 26

FONT_NAME = "Aptos Narrow"  # misma tipografia del archivo original del cliente

wb = openpyxl.Workbook()

# ---------- estilos compartidos ----------
TITLE = Font(name=FONT_NAME, size=17, bold=True, color="1F3864")
SUBTLE = Font(name=FONT_NAME, size=9.5, italic=True, color="7F7F7F")
LABELF = Font(name=FONT_NAME, size=10.5, bold=True)
HEAD = Font(name=FONT_NAME, size=9.5, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="3A5F8A")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
NUM = Font(name=FONT_NAME, size=10)
POS = Font(name=FONT_NAME, size=10, color="0A7A2E", bold=True)
NEG = Font(name=FONT_NAME, size=10, color="C0392B", bold=True)
SECTION = Font(name=FONT_NAME, size=12, bold=True, color="1F3864")
NOTEF = Font(name=FONT_NAME, size=9.5, italic=True, color="595959")
NEWF = Font(name=FONT_NAME, size=9.5, italic=True, color="1F6FB2")
thin = Side(style="thin", color="D9D9D9")
WRAP = Alignment(wrap_text=True, vertical="top")

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center" if c > start_col else "left", vertical="center")

def pct_cell(ws, row, col, value, is_new=False):
    if is_new:
        cell = ws.cell(row=row, column=col, value="Nuevo (post S30)")
        cell.font = NEWF
        cell.alignment = Alignment(horizontal="right")
        return cell
    cell = ws.cell(row=row, column=col, value=(value / 100 if value is not None else None))
    cell.number_format = "+0.0%;-0.0%;0.0%"
    cell.font = POS if (value or 0) >= 0 else NEG
    return cell

# ============================================================
# Hoja 0: INICIO (portada / orientacion para el cliente)
# ============================================================
cov = wb.active
cov.title = "INICIO"
cov.sheet_view.showGridLines = False
cov.sheet_properties.tabColor = "1F3864"

cov["B2"] = "BESHOS"
cov["B2"].font = Font(name=FONT_NAME, size=26, bold=True, color="1F3864")
cov["B3"] = "Seguimiento comercial retail"
cov["B3"].font = Font(name=FONT_NAME, size=14, color="595959")

cov["B5"] = "Periodo cubierto"
cov["B5"].font = LABELF
cov["C5"] = "Semana ISO 1 a 33, 2026"
cov["C5"].font = NUM
cov["B6"] = "Toma de cuenta"
cov["B6"].font = LABELF
cov["C6"] = "22-06-2026  (semana ISO 26)"
cov["C6"].font = NUM
cov["B7"] = "Metrica"
cov["B7"].font = LABELF
cov["C7"] = "Unidades vendidas (Qty)"
cov["C7"].font = NUM
cov["B8"] = "Generado"
cov["B8"].font = LABELF
cov["C8"] = "25-08-2026"
cov["C8"].font = NUM

cov["B10"] = "Contenido"
cov["B10"].font = SECTION
guide = [
    ("RESUMEN RETAIL", "Impacto antes / desde la toma de cuenta por retail, en numeros."),
    ("EVOLUCIÓN GESTIÓN", "Evolucion semanal desde la toma de cuenta (sem. 26), un grafico por retail."),
    ("DETALLE SKU", "Que productos explican la venta de cada retail y como vienen evolucionando."),
    ("DETALLE SALAS", "Que locales concentran la venta de cada retail."),
]
r = 11
for name, desc in guide:
    cov.cell(row=r, column=2, value=name).font = LABELF
    cov.cell(row=r, column=3, value=desc).font = NUM
    cov.cell(row=r, column=3).alignment = WRAP
    cov.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    cov.row_dimensions[r].height = 28
    r += 1

r += 1
cov.cell(row=r, column=2,
         value="Fuente: hojas VTA_WALMART / VTA_JUMBO / VTA_SMU / VTA_TOTTUS del archivo base ANALISIS_BESHOS.xlsx.").font = SUBTLE

col_widths_cov = {1: 3, 2: 20, 3: 22, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14}
for c, w in col_widths_cov.items():
    cov.column_dimensions[get_column_letter(c)].width = w

# ============================================================
# Hoja 1: RESUMEN RETAIL -- solo numeros (KPI + tabla fuente), sin graficos
# ============================================================
ws = wb.create_sheet("RESUMEN RETAIL")
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = "3A5F8A"

ws["A1"] = "BESHOS \u2014 Resumen gesti\u00f3n retail"
ws["A1"].font = TITLE
ws.merge_cells("A1:H1")

ws["A2"] = ("Unidades vendidas \u00b7 semana ISO 1 a 33, 2026 \u00b7 Toma de cuenta: 22-06-2026 (semana ISO 26). "
            "SMU y Tottus solo tienen data reportada desde semana 19.")
ws["A2"].font = SUBTLE
ws.merge_cells("A2:H2")
ws.row_dimensions[2].height = 26
ws["A2"].alignment = WRAP

r = 4
ws.cell(row=r, column=1, value="Impacto de gesti\u00f3n \u2014 promedio semanal (unidades)").font = SECTION
r += 1
headers = ["Retail", "Sem. antes RT", "Prom/sem antes RT", "Sem. desde RT", "Prom/sem desde RT", "Variaci\u00f3n"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=r, column=c, value=h)
style_header_row(ws, r, len(headers))
kpi_header_row = r
r += 1

for ret in RETAILERS:
    pp = R[ret]["pre_post_takeover"]
    ws.cell(row=r, column=1, value=LABELS[ret]).font = LABELF
    ws.cell(row=r, column=2, value=pp["pre_weeks"]).font = NUM
    ws.cell(row=r, column=3, value=pp["pre_avg_per_week"]).font = NUM
    ws.cell(row=r, column=3).number_format = "#,##0"
    ws.cell(row=r, column=4, value=pp["post_weeks"]).font = NUM
    ws.cell(row=r, column=5, value=pp["post_avg_per_week"]).font = NUM
    ws.cell(row=r, column=5).number_format = "#,##0"
    pct_cell(ws, r, 6, pp["pct_change"])
    if (r - kpi_header_row) % 2 == 0:
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = ALT_FILL
    r += 1

r += 1
divider_row = r
for c in range(1, 9):
    ws.cell(row=divider_row, column=c).fill = PatternFill("solid", fgColor="1F3864")
ws.row_dimensions[divider_row].height = 4
r += 2

ws.cell(row=r, column=1, value="Serie semanal (unidades) \u2014 detalle S1\u2013S33").font = SECTION
r += 1
ws.cell(row=r, column=1,
        value="Los gr\u00e1ficos de esta serie est\u00e1n en la hoja EVOLUCI\u00d3N GESTI\u00d3N.").font = NOTEF
r += 1
weeks_header_row = r
ws.cell(row=r, column=1, value="Semana")
series_cols = {}
col = 2
for ret in RETAILERS:
    ws.cell(row=r, column=col, value=LABELS[ret])
    series_cols[ret] = col
    col += 1
style_header_row(ws, r, col - 1)
r += 1
first_data_row = r

week_units = {ret: {p["iso_week"]: p["units"] for p in R[ret]["weekly_series"]} for ret in RETAILERS}

for wk in range(1, 34):
    ws.cell(row=r, column=1, value=f"S{wk}").font = NUM
    for ret in RETAILERS:
        v = week_units[ret].get(wk)
        c = ws.cell(row=r, column=series_cols[ret], value=v)
        c.font = NUM
        c.number_format = "#,##0"
    r += 1
last_data_row = r - 1

# Colapsar la tabla de detalle semanal: en esta hoja el foco son los numeros
# resumen; el detalle completo queda a un clic de distancia ("+" del outline).
ws.sheet_properties.outlinePr.summaryBelow = False
for rr in range(first_data_row, last_data_row + 1):
    ws.row_dimensions[rr].outlineLevel = 1
    ws.row_dimensions[rr].hidden = True

ws.cell(row=last_data_row + 2, column=1,
        value="Fuente: VTA_WALMART / VTA_JUMBO / VTA_SMU / VTA_TOTTUS. Detalle completo en analisis_beshos_retail.json.").font = SUBTLE

col_widths = {1: 22, 2: 15, 3: 15, 4: 15, 5: 15, 6: 15}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

# ============================================================
# Hoja 2: EVOLUCIÓN GESTIÓN -- graficos, estilo ejecutivo, solo desde sem.26
# ============================================================
wg = wb.create_sheet("EVOLUCIÓN GESTIÓN")
wg.sheet_view.showGridLines = False
wg.sheet_properties.tabColor = "1F3864"

wg["A1"] = "Evoluci\u00f3n de la gesti\u00f3n \u2014 desde la toma de cuenta"
wg["A1"].font = TITLE
wg.merge_cells("A1:H1")
wg["A2"] = "Unidades vendidas por semana \u00b7 semana ISO 26 a 33, 2026 (22-06-2026 en adelante)."
wg["A2"].font = SUBTLE
wg.merge_cells("A2:H2")

# ---- bloque de datos auxiliar (colapsado), S26-S33 unicamente ----
data_row0 = 120  # bien debajo de las 5 tarjetas para que las filas ocultas no interfieran con ningun grafico
wg.cell(row=data_row0, column=1, value="Semana")
gest_cols = {}
gc = 2
for ret in RETAILERS:
    wg.cell(row=data_row0, column=gc, value=LABELS[ret])
    gest_cols[ret] = gc
    gc += 1
gest_first = data_row0 + 1
for i, wk in enumerate(range(BASE_WEEK, 34)):
    rr = gest_first + i
    wg.cell(row=rr, column=1, value=f"S{wk}")
    for ret in RETAILERS:
        wg.cell(row=rr, column=gest_cols[ret], value=week_units[ret].get(wk))
gest_last = gest_first + (33 - BASE_WEEK)
for rr in range(data_row0, gest_last + 1):
    wg.row_dimensions[rr].outlineLevel = 1
    wg.row_dimensions[rr].hidden = True

# ---- tarjetas ejecutivas: nombre + variacion + minigrafico de tendencia ----
ACCENT = "1F3864"       # azul corporativo -- linea de tendencia
ACCENT_FILL = "D9E2F3"  # relleno suave bajo la linea
CARD_BORDER = Side(style="thin", color="D9D9D9")
CARD_BOX = Border(left=CARD_BORDER, right=CARD_BORDER, top=CARD_BORDER, bottom=CARD_BORDER)

def badge(ws_, row, col, pct):
    up = (pct or 0) >= 0
    fill = PatternFill("solid", fgColor="E3F3E8" if up else "FBE7E5")
    color = "0A7A2E" if up else "C0392B"
    cell = ws_.cell(row=row, column=col, value=(pct / 100 if pct is not None else None))
    cell.number_format = "+0.0%;-0.0%;0.0%"
    cell.font = Font(name=FONT_NAME, size=13, bold=True, color=color)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell

CHART_W_CM = 17.0
CHART_H_CM = 5.2

def trend_chart(ret, anchor):
    c_idx = gest_cols[ret]
    chart = LineChart()
    chart.title = None
    chart.legend = None
    chart.visible_cells_only = False  # las filas fuente estan colapsadas (outline)
    chart.height = CHART_H_CM
    chart.width = CHART_W_CM
    chart.y_axis.delete = True
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorTickMark = "none"
    chart.x_axis.minorTickMark = "none"
    chart.x_axis.tickLblPos = "low"
    cats = Reference(wg, min_col=1, min_row=gest_first, max_row=gest_last)
    data_ref = Reference(wg, min_col=c_idx, min_row=gest_first, max_row=gest_last)
    chart.add_data(data_ref, titles_from_data=False)
    s = chart.series[-1]
    s.marker = Marker(symbol="circle", size=6)
    s.smooth = False
    s.graphicalProperties.line.width = 26000
    s.graphicalProperties.line.solidFill = ACCENT
    s.marker.graphicalProperties.solidFill = ACCENT
    s.marker.graphicalProperties.line.solidFill = ACCENT
    # etiqueta solo en el ultimo punto: el numero que importa hoy
    last_idx = gest_last - gest_first
    dlbl = DataLabel(idx=last_idx)
    dlbl.showVal = True
    dlbl.showSerName = False
    dlbl.showCatName = False
    dlbl.showLegendKey = False
    dlbl.numFmt = "#,##0"
    s.dLbls = DataLabelList(
        dLbl=[dlbl],
        showVal=False, showCatName=False, showSerName=False,
        showLegendKey=False, showPercent=False, showBubbleSize=False,
    )
    chart.set_categories(cats)
    wg.add_chart(chart, anchor)
    return chart

# Una sola columna de tarjetas apiladas: evita cualquier riesgo de que el
# ancho del grafico (en cm) no calce con el ancho real de las columnas de
# Excel (eso fue lo que produjo el desborde/superposicion de la version
# anterior). Alto de fila por defecto ~0.53cm -> CHART_H_CM=5.2cm ~= 10 filas;
# se deja un salto de 14 filas por tarjeta para que sobre margen de sobra.
ROWS_PER_CARD = 14
card_start_row = 4

for i, ret in enumerate(RETAILERS):
    row_num = card_start_row + i * ROWS_PER_CARD
    wg.cell(row=row_num, column=2, value=LABELS[ret]).font = Font(name=FONT_NAME, size=13, bold=True, color="1F3864")
    pct = R[ret]["pre_post_takeover"]["pct_change"]
    badge(wg, row_num, 4, pct)
    wg.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=5)
    wg.cell(row=row_num + 1, column=2,
            value=f"unid./semana \u00b7 prom. {R[ret]['pre_post_takeover']['post_avg_per_week']:,.0f} desde S26".replace(",", ".")
            ).font = NOTEF
    trend_chart(ret, f"B{row_num + 2}")

last_card_row = card_start_row + (len(RETAILERS) - 1) * ROWS_PER_CARD

col_widths_g = {1: 3, 2: 18, 3: 14, 4: 12, 5: 12}
for c, w in col_widths_g.items():
    wg.column_dimensions[get_column_letter(c)].width = w

wg.cell(row=last_card_row + 11, column=1,
        value="Fuente: VTA_WALMART / VTA_JUMBO / VTA_SMU / VTA_TOTTUS. Walmart considera solo el surtido activo de siempre "
              "(4 SKU). Detalle completo en analisis_beshos_retail.json.").font = SUBTLE
wg.merge_cells("A34:H34")
wg.row_dimensions[34].height = 26
wg["A34"].alignment = WRAP

# ============================================================
# Hoja 3: DETALLE SKU
# ============================================================
ws2 = wb.create_sheet("DETALLE SKU")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = "5B8DB8"
ws2["A1"] = "Comportamiento por SKU \u2014 antes / desde toma de cuenta"
ws2["A1"].font = TITLE
ws2.merge_cells("A1:G1")
ws2["A2"] = "Top SKU por volumen 2026 (unidades). Var % compara promedio semanal antes vs. desde sem. ISO 26."
ws2["A2"].font = SUBTLE
ws2.merge_cells("A2:G2")
ws2.freeze_panes = "A4"

col_widths2 = {1: 12, 2: 38, 3: 13, 4: 10, 5: 15, 6: 15, 7: 15}
for c, w in col_widths2.items():
    ws2.column_dimensions[get_column_letter(c)].width = w

r = 4
headers_sku = ["SKU", "Producto", "Unid. totales", "% part.", "Prom/sem antes", "Prom/sem desde", "Variaci\u00f3n"]
for ret in RETAILERS:
    label = LABELS[ret]
    rows = R[ret]["top_skus"][:6]
    n_skus = R[ret]["sku_count"]
    ws2.cell(row=r, column=1, value=f"{label}  ({n_skus} SKU activos)").font = SECTION
    r += 1
    for c, h in enumerate(headers_sku, start=1):
        ws2.cell(row=r, column=c, value=h)
    style_header_row(ws2, r, len(headers_sku))
    r += 1
    for row_i, s in enumerate(rows):
        ws2.cell(row=r, column=1, value=s["sku"]).font = NUM
        ws2.cell(row=r, column=2, value=s["desc"]).font = NUM
        ws2.cell(row=r, column=3, value=s["total_units"]).font = NUM
        ws2.cell(row=r, column=3).number_format = "#,##0"
        ws2.cell(row=r, column=4, value=(s["share_pct"] / 100 if s["share_pct"] is not None else None)).number_format = "0.0%"
        ws2.cell(row=r, column=4).font = NUM
        ws2.cell(row=r, column=5, value=s["pre_avg_per_week"]).font = NUM
        ws2.cell(row=r, column=5).number_format = "#,##0.0"
        ws2.cell(row=r, column=6, value=s["post_avg_per_week"]).font = NUM
        ws2.cell(row=r, column=6).number_format = "#,##0.0"
        pct_cell(ws2, r, 7, s["pct_change"], is_new=(s["pre_avg_per_week"] is None))
        if row_i % 2 == 1:
            for c in range(1, 8):
                ws2.cell(row=r, column=c).fill = ALT_FILL
        r += 1
    r += 1

# ============================================================
# Hoja 4: DETALLE SALAS
# ============================================================
ws3 = wb.create_sheet("DETALLE SALAS")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = "5B8DB8"
ws3["A1"] = "Comportamiento por sala \u2014 antes / desde toma de cuenta"
ws3["A1"].font = TITLE
ws3.merge_cells("A1:G1")
ws3["A2"] = "Top salas por volumen 2026 (unidades). Var % compara promedio semanal antes vs. desde sem. ISO 26."
ws3["A2"].font = SUBTLE
ws3.merge_cells("A2:G2")
ws3.freeze_panes = "A4"

for c, w in col_widths2.items():
    ws3.column_dimensions[get_column_letter(c)].width = w

r = 4
headers_store = ["C\u00f3d.", "Sala", "Unid. totales", "% part.", "Prom/sem antes", "Prom/sem desde", "Variaci\u00f3n"]
for ret in RETAILERS:
    label = LABELS[ret]
    n_stores = R[ret]["store_count"]
    ws3.cell(row=r, column=1, value=f"{label}  ({n_stores} salas)").font = SECTION
    r += 1
    for c, h in enumerate(headers_store, start=1):
        ws3.cell(row=r, column=c, value=h)
    style_header_row(ws3, r, len(headers_store))
    r += 1
    for row_i, s in enumerate(R[ret]["top_stores"][:5]):
        ws3.cell(row=r, column=1, value=s["store"]).font = NUM
        ws3.cell(row=r, column=2, value=s["store_name"]).font = NUM
        ws3.cell(row=r, column=3, value=s["total_units"]).font = NUM
        ws3.cell(row=r, column=3).number_format = "#,##0"
        ws3.cell(row=r, column=4, value=(s["share_pct"] / 100 if s["share_pct"] is not None else None)).number_format = "0.0%"
        ws3.cell(row=r, column=4).font = NUM
        ws3.cell(row=r, column=5, value=s["pre_avg_per_week"]).font = NUM
        ws3.cell(row=r, column=5).number_format = "#,##0.0"
        ws3.cell(row=r, column=6, value=s["post_avg_per_week"]).font = NUM
        ws3.cell(row=r, column=6).number_format = "#,##0.0"
        pct_cell(ws3, r, 7, s["pct_change"], is_new=(s["pre_avg_per_week"] is None))
        if row_i % 2 == 1:
            for c in range(1, 8):
                ws3.cell(row=r, column=c).fill = ALT_FILL
        r += 1
    r += 1

wb.active = 0
wb.save(OUT_PATH)
print("saved:", OUT_PATH)
