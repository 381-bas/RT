# -*- coding: utf-8 -*-
import openpyxl, json, datetime
from collections import defaultdict

SRC = r"C:\Users\basti\Desktop\RT\CLIENTE_GG\ANÁLISIS_BESHOS.xlsx"

BASE_YEAR, BASE_WEEK = 2026, 26   # toma de cuenta

# Walmart: el surtido real y consistente de BESHOS son estos 4 SKU. Desde
# sem.30 el reporte de Walmart se contamino con ~13 codigos de otras lineas
# (nectares, te helado) que no son de este cliente -- se descartan por
# completo en el ingreso de datos, no solo en la presentacion: no se
# consideran en ninguna cifra ni grafico del informe.
WALMART_ACTIVE_SKUS = {700419, 700420, 700421, 5064681}

wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

# retailer -> week -> qty
weekly = defaultdict(lambda: defaultdict(float))
# retailer -> (sku, desc) -> week -> qty
sku_weekly = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
# retailer -> (store, name) -> week -> qty
store_weekly = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

def add(retailer, week, qty, sku=None, sku_desc=None, store=None, store_name=None):
    weekly[retailer][week] += qty
    if sku is not None:
        sku_weekly[retailer][(sku, sku_desc)][week] += qty
    if store is not None:
        store_weekly[retailer][(store, store_name)][week] += qty

def y2026(ano, iso):
    try:
        ano = int(ano); iso = int(iso)
    except Exception:
        return None
    if ano != BASE_YEAR:
        return None
    return iso

# ---- WALMART ----
ws = wb['VTA_WALMART']
for i, row in enumerate(ws.iter_rows(min_col=1, max_col=14, values_only=True)):
    if i == 0:
        continue
    sku, desc, store, store_name, qty, ano, iso = row[0], row[1], row[2], row[3], row[4], row[10], row[13]
    if qty is None or sku not in WALMART_ACTIVE_SKUS:
        continue
    w = y2026(ano, iso)
    if w is None:
        continue
    add('WALMART', w, float(qty), sku, desc, store, store_name)

# ---- JUMBO / SANTA ISABEL ----
ws = wb['VTA_JUMBO']
for i, row in enumerate(ws.iter_rows(min_col=1, max_col=29, values_only=True)):
    if i == 0:
        continue
    prod, desc, store, store_name, qty, ano, iso, sub = row[1], row[3], row[9], row[11], row[17], row[23], row[28], row[27]
    if qty is None or sub not in ('JUMBO', 'SANTA ISABEL'):
        continue
    w = y2026(ano, iso)
    if w is None:
        continue
    add(sub, w, float(qty), prod, desc, store, store_name)

# ---- SMU ----
ws = wb['VTA_SMU']
for i, row in enumerate(ws.iter_rows(min_col=1, max_col=24, values_only=True)):
    if i == 0:
        continue
    prod, desc, store, store_name, qty, ano, iso = row[1], row[3], row[7], row[9], row[13], row[20], row[23]
    if qty is None:
        continue
    w = y2026(ano, iso)
    if w is None:
        continue
    add('SMU', w, float(qty), prod, desc, store, store_name)

# ---- TOTTUS ----
ws = wb['VTA_TOTTUS']
for i, row in enumerate(ws.iter_rows(min_col=1, max_col=25, values_only=True)):
    if i == 0:
        continue
    prod, desc, store, store_name, qty, ano, iso = row[7], row[8], row[10], row[11], row[15], row[21], row[24]
    if qty is None:
        continue
    w = y2026(ano, iso)
    if w is None:
        continue
    add('TOTTUS', w, float(qty), prod, desc, store, store_name)

wb.close()

RETAILERS = ['WALMART', 'JUMBO', 'SANTA ISABEL', 'SMU', 'TOTTUS']

def week_to_monday(w):
    return datetime.date.fromisocalendar(BASE_YEAR, w, 1)

def pre_post(week_qty):
    """avg/week before BASE_WEEK vs from BASE_WEEK on."""
    pre = [v for w, v in week_qty.items() if w < BASE_WEEK]
    post = [v for w, v in week_qty.items() if w >= BASE_WEEK]
    pre_avg = sum(pre) / len(pre) if pre else None
    post_avg = sum(post) / len(post) if post else None
    pct = None
    if pre_avg and post_avg is not None:
        pct = round((post_avg - pre_avg) / pre_avg * 100, 1)
    return {
        "pre_weeks": len(pre), "pre_avg_per_week": round(pre_avg, 1) if pre_avg else None,
        "post_weeks": len(post), "post_avg_per_week": round(post_avg, 1) if post_avg else None,
        "pct_change": pct,
    }

result = {}
for ret in RETAILERS:
    wk = weekly[ret]
    weeks_sorted = sorted(wk.keys())
    series = [{"iso_week": w, "week_start": week_to_monday(w).isoformat(), "units": round(wk[w], 1)} for w in weeks_sorted]
    total_units = round(sum(wk.values()), 1)

    # top SKUs by total volume
    sku_rows = []
    for (sku, desc), wq in sku_weekly[ret].items():
        total = sum(wq.values())
        pp = pre_post(wq)
        sku_rows.append({
            "sku": sku, "desc": desc, "total_units": round(total, 1),
            "share_pct": round(total / total_units * 100, 1) if total_units else None,
            **pp
        })
    sku_rows.sort(key=lambda r: -r["total_units"])

    # top stores by total volume
    store_rows = []
    for (store, sname), wq in store_weekly[ret].items():
        total = sum(wq.values())
        pp = pre_post(wq)
        store_rows.append({
            "store": store, "store_name": sname, "total_units": round(total, 1),
            "share_pct": round(total / total_units * 100, 1) if total_units else None,
            **pp
        })
    store_rows.sort(key=lambda r: -r["total_units"])

    result[ret] = {
        "weekly_series": series,
        "total_units_2026": total_units,
        "pre_post_takeover": pre_post(wk),
        "sku_count": len(sku_rows),
        "store_count": len(store_rows),
        "top_skus": sku_rows[:8],
        "bottom_movers_skus": sorted([r for r in sku_rows if r["pre_avg_per_week"] and r["pre_weeks"] >= 4 and r["post_weeks"] >= 3],
                                      key=lambda r: r["pct_change"])[:5],
        "top_stores": store_rows[:8],
        "bottom_movers_stores": sorted([r for r in store_rows if r["pre_avg_per_week"] and r["pre_weeks"] >= 4 and r["post_weeks"] >= 3],
                                        key=lambda r: r["pct_change"])[:5],
    }

out = {
    "source_file": SRC,
    "generated_at": datetime.datetime.now().isoformat(),
    "metric": "units (Qty)",
    "period_note": "Todas las semanas ISO 2026 con data disponible (S1 a S33).",
    "baseline_note": "Toma de cuenta: 22-06-2026 = ISO semana 26, 2026. pre_post_takeover compara promedio semanal ANTES (S1-25) vs DESDE (S26-33).",
    "walmart_note": (
        "WALMART se filtra desde el ingreso de datos a los 3 SKU realmente activos de BESHOS "
        "(700420, 700421, 5064681). El resto de codigos que aparecen en VTA_WALMART desde sem.30 "
        "corresponden a otras lineas de producto (nectares, te helado) y se descartan por completo: "
        "no se consideran en ninguna cifra ni grafico de este informe."
    ),
    "retailers": result,
}

out_path = r"C:\Users\basti\Desktop\RT\CLIENTE_GG\analisis_beshos_retail.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

for ret in RETAILERS:
    d = result[ret]
    print(ret, "weeks:", len(d["weekly_series"]), "skus:", d["sku_count"], "stores:", d["store_count"],
          "pre/post:", d["pre_post_takeover"]["pre_avg_per_week"], "->", d["pre_post_takeover"]["post_avg_per_week"],
          d["pre_post_takeover"]["pct_change"])
print("saved:", out_path)
