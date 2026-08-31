#!/usr/bin/env python3
"""
Procesador de análisis de cierre BESHOS — Opción 2C.2A

Crea propuesta alternativa de distribución de costos:
- Propuesta 1: Cobrar a todos (entrada)
- Propuesta 2C.2A: Regalar SMU + Tottus, distribuir costo con ponderación estratégica
  * Walmart 42% (líder)
  * Jumbo 33% (segundo pilar)
  * Santa Isabel 25% (equilibrado)
  * SMU + Tottus 0% (plan piloto gratis)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json

def crear_propuesta_2c2a(input_file, output_file):
    """Genera archivo con Propuesta 2C.2A"""

    # Cargar archivo
    wb = openpyxl.load_workbook(input_file)
    ws = wb['RESUMEN']

    # Datos de entrada (Opción 1)
    opciones = {
        'CENCOSUD_JUMBO': {'salas': 58, 'costo_opc1': 756668, 'tarifa_opc1': 13046},
        'CENCOSUD_SI': {'salas': 88, 'costo_opc1': 716496, 'tarifa_opc1': 8142},
        'SMU': {'salas': 16, 'costo_opc1': 114064, 'tarifa_opc1': 7129},
        'TOTTUS': {'salas': 27, 'costo_opc1': 199700.1, 'tarifa_opc1': 7396.3},
        'WALMART': {'salas': 93, 'costo_opc1': 1142412, 'tarifa_opc1': 12284},
    }

    # Costo total a mantener
    costo_total = 2929340.1
    salas_pagadoras = 239  # Walmart (93) + Jumbo (58) + SI (88)

    # Ponderación 2C.2A
    ponderacion = {
        'WALMART': 0.42,
        'CENCOSUD_JUMBO': 0.33,
        'CENCOSUD_SI': 0.25,
        'SMU': 0.0,
        'TOTTUS': 0.0,
    }

    # Calcular Propuesta 2C.2A
    propuesta_2c2a = {}
    for cadena, peso in ponderacion.items():
        salas = opciones[cadena]['salas']
        if peso > 0:
            costo = costo_total * peso
            tarifa = costo / salas
            propuesta_2c2a[cadena] = {
                'salas': salas,
                'costo': round(costo, 2),
                'tarifa': round(tarifa, 2),
                'cambio_pct': round(((tarifa - opciones[cadena]['tarifa_opc1']) / opciones[cadena]['tarifa_opc1']) * 100, 1)
            }
        else:
            propuesta_2c2a[cadena] = {
                'salas': salas,
                'costo': 0,
                'tarifa': 0,
                'cambio_pct': -100  # Ahorro del 100%
            }

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    currency_format = '_($* #,##0.0_);_($* (#,##0.0);_($* "-"??_);_(@_)'

    # Escribir Propuesta 2C.2A en la hoja
    row_start = 11  # Después de Opción 2 placeholder
    ws[f'A{row_start}'] = "OPCIÓN 2C.2A (APROBADA)"
    ws[f'A{row_start}'].font = Font(bold=True, size=12, color="C00000")

    row_start += 1
    headers = ['CADENA', 'FORMATO', 'SALAS', 'COSTO EMPRESA', 'COSTO POR SALA']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_start, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Datos Propuesta 2C.2A
    cadenas_orden = ['CENCOSUD_JUMBO', 'CENCOSUD_SI', 'SMU', 'TOTTUS', 'WALMART']
    row_start += 1
    total_costo = 0
    total_salas = 0

    for idx, cadena in enumerate(cadenas_orden):
        row = row_start + idx
        data = propuesta_2c2a[cadena]

        # Nombre cadena
        if 'JUMBO' in cadena:
            nombre = 'CENCOSUD'
            formato = 'JUMBO'
        elif 'SI' in cadena:
            nombre = 'CENCOSUD'
            formato = 'SANTA ISABEL'
        elif 'SMU' in cadena:
            nombre = 'SMU'
            formato = 'UNIMARC'
        elif 'TOTTUS' in cadena:
            nombre = 'TOTTUS'
            formato = 'TOTTUS'
        elif 'WALMART' in cadena:
            nombre = 'WALMART'
            formato = 'WALMART'

        ws[f'A{row}'].value = nombre
        ws[f'B{row}'].value = formato
        ws[f'C{row}'].value = data['salas']
        ws[f'D{row}'].value = data['costo']
        ws[f'E{row}'].value = data['tarifa']

        # Estilos
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            if col in [4, 5]:
                cell.number_format = currency_format

        # Alineación izquierda para texto
        ws[f'A{row}'].alignment = Alignment(horizontal='left')
        ws[f'B{row}'].alignment = Alignment(horizontal='left')
        ws[f'C{row}'].alignment = Alignment(horizontal='center')

        total_costo += data['costo']
        total_salas += data['salas']

    # Total
    row_total = row_start + 5
    ws[f'A{row_total}'].value = 'TOTAL'
    ws[f'C{row_total}'].value = total_salas
    ws[f'D{row_total}'].value = round(total_costo, 2)
    ws[f'E{row_total}'].value = round(total_costo / total_salas, 2) if total_salas > 0 else 0

    for col in range(1, 6):
        cell = ws.cell(row=row_total, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        if col in [4, 5]:
            cell.number_format = currency_format
        if col in [1, 2]:
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = Alignment(horizontal='right')

    # Guardار
    wb.save(output_file)
    print(f"✓ Archivo generado: {output_file}")

    # Retornar datos para JSON
    return {
        'propuesta': '2C.2A',
        'descripcion': 'Regalar SMU+Tottus, distribuir costo: Walmart 42%, Jumbo 33%, SI 25%',
        'costo_total_empresa': costo_total,
        'salas_totales': total_salas,
        'salas_pagadoras': total_salas,
        'salas_gratis': 43,
        'detalle': {
            cadena: {
                'salas': data['salas'],
                'costo': data['costo'],
                'tarifa_por_sala': data['tarifa'],
                'cambio_porcentaje': data['cambio_pct']
            }
            for cadena, data in propuesta_2c2a.items()
        }
    }

if __name__ == '__main__':
    input_f = 'Analisis_cierre.xlsx'
    output_f = 'Analisis_cierre_cliente.xlsx'

    print(f"Procesando {input_f}...")
    datos = crear_propuesta_2c2a(input_f, output_f)

    # Guardar trazabilidad
    with open('_generated/trazabilidad_cierre_31_08.json', 'w', encoding='utf-8') as f:
        json.dump({
            'fecha_ejecucion': datetime.now().isoformat(),
            'tarea': 'Cierre_Piloto_31_08',
            'entrada': input_f,
            'salida': output_f,
            'propuesta': datos,
            'ejecutado_por': 'claude'
        }, f, indent=2, ensure_ascii=False)

    print("✓ Trazabilidad guardada")
